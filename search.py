from googleapiclient.discovery import build
import openpyxl

# created to find websites of some companies
# requires api key and custom search ID
# more details: https://developers.google.com/custom-search/v1/overview
my_api_key = ''
my_cse_id = ''


# get's the list of excluded websites from text file
def get_excluded():
    with open('excluded_sites.txt', 'r') as file:
        return file.readline()


# google search api call
# parameter list: https://developers.google.com/custom-search/v1/cse/list
def google_search(search_term, api_key, cse_id, **kwargs):
    service = build("customsearch", "v1", developerKey=api_key)
    res = service.cse().list(q=search_term, cx=cse_id, **kwargs).execute()
    # returns the search results
    return res['items']


# reading the country iso codes from separate file
# needed to make searches location based
# returns dictionary {CountryID: [country name, country iso2 code]}
def get_country_data():
    wb_data = openpyxl.load_workbook('data.xlsx')
    ws_country = wb_data['Country']

    country = {}

    for row in range(2, ws_country.max_row + 1):
        country_name = ws_country.cell(row=row, column=2).value
        country_iso2 = ws_country.cell(row=row, column=3).value

        country[ws_country.cell(row=row, column=1).value] = [country_name, country_iso2]

    return country


# fetch the google search result
# get's the company information from excel file, process the results and puts them in the file
def get_values(excel_file, sheet_name, company_name_column, city_name_column, country_column, result_column):

    excluded = get_excluded()
    country_list = get_country_data()

    # process first 100 rows
    # to process all rows in column use: for row in range(2, sheet_name.max_row + 1)):
    for row in range(2, 3):

        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb[sheet_name]

        company_name = worksheet.cell(row=row, column=company_name_column).value
        city_name = worksheet.cell(row=row, column=city_name_column).value

        # in case there is NULL as company name, will skip this kind of cases
        if company_name != 'NULL':

            # if city name is not provided
            if city_name is not None:
                search_string = ' '.join([company_name, city_name])
            else:
                search_string = company_name

            # if in the source file there is no country information, assume US
            if country_column != 0:
                search_country = country_list[worksheet.cell(row=row, column=country_column).value][1]
            else:
                search_country = 'US'

            # sometimes may get no results and KeyError. Probably due to excluded sites filter.
            # in such cases it's calling the API without this filter
            try:
                worksheet.cell(row=row, column=result_column).value = google_search(search_string, my_api_key,
                                                                                    my_cse_id,
                                                                                    num=1, siteSearch=excluded,
                                                                                    siteSearchFilter='e',
                                                                                    gl=search_country)[0]['link']
            except KeyError:
                worksheet.cell(row=row, column=result_column).value = google_search(search_string, my_api_key,
                                                                                    my_cse_id, num=1,
                                                                                    gl=search_country)[0]['link']

            wb.save(excel_file)

        wb.save(excel_file)


get_values('without Websites US CAN.xlsx', 'without Websites US CAN', 2, 5, 0, 6)


# print(get_secret()[0])
